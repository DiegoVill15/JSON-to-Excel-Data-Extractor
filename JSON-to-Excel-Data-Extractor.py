#Programa para Obtener datos .json y guardarlos en Excel
#@Author: Diego Villalobos
#@date:04/12/2023

import requests
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Constantes configurables
URL_BASE = "http://paginadeprueba.com/BD/"
HEADERS = ["ID", "Nombre", "Apellido paterno", "Apellido Materno", "Género", "Fecha de Nacimiento", "Email", "Teléfono"]
FILENAME = "Datos_cliente.xlsx"

#Obtener datos por medio de request y convertirlos en .json para leerlos y copiarlos en excel
def get_data_from_api(id_value):
    url = f"{URL_BASE}{id_value}"
    response = requests.get(url)
    if response.status_code == 200:
        return json.loads(response.text)
    else:
        print(f"Error al obtener datos para el ID {id_value}: {response.status_code}")
        return {'id': id_value, 'error': 'Sin datos encontrados'}


#Creación de documento en excel y guardado de datos.
def save_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos_cliente"
    ws.append(HEADERS)
    
    # Estilo de celda roja (no encontrado)
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for item in data:
        if 'error' in item:
            ws.append([item['id'], item['error']])
            for cell in ws[ws.max_row]:
                cell.fill = red_fill
        else:
            row = [
                item.get('id', ''),
                item.get('name', ''),
                item.get('lastName1', ''),
                item.get('lastName2', ''),
                item.get('gender', ''),
                item.get('birthDate', ''),
                item.get('email', ''),
                item.get('phone', ''),
            ]
            ws.append(row)

    wb.save(filename=FILENAME)

# Main
entries = [('11', '111111'), ('22', '2222222222'), ('333', '333333')]
results = []

for id_type, id_value in entries:
    applicant_data = get_data_from_api(id_value)
    results.append(applicant_data)

print("*** GENERANDO DOCUMENTO ***")
save_to_excel(results)
print("*** DOCUMENTO GENERADO ***")
